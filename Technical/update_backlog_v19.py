"""
Update RentHack_Product_Backlog_v18.xlsx with the 2026-06 accuracy & UX audit
work (calc-engine P0 fixes, display/export corrections, onboarding fixes),
then save as v19.

Note: v18 (308 rows) superseded v17 — CLAUDE.md previously pointed at v17;
this run rebases on v18 and CLAUDE.md is updated to v19.

Run from project root:
  python Technical/update_backlog_v19.py
"""
import sys, io, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from copy import copy
from pathlib import Path

PROJ = Path(__file__).parent.parent
SRC = PROJ / "RentHack_Product_Backlog_v18.xlsx"
DST = PROJ / "RentHack_Product_Backlog_v19.xlsx"

# ─── 1. Status updates for existing rows ────────────────────────────────────
STATUS_UPDATES = []  # audit work is all new — no existing IDs covered it

# ─── 2. New rows to append ──────────────────────────────────────────────────
#   ID, Epic, Feature/Task, Priority, Status, Effort, User Story,
#   Acceptance Criteria/Build Notes, Why It Matters, Dependencies
NEW_ROWS = [
    # ─── 2026-06 Accuracy Audit — calc engine (P0) ─────────────────────────
    ("BACK-095", "Financial Model",
     "Honor zero-value inputs — 0% down (VA), 0% rate, 0% tax bracket",
     "P0 – Foundation", "Done (PROD)", "S",
     "As a VA buyer modeling 0% down, I want the calculator to actually use 0% instead of silently substituting 25%.",
     "◇ Build complete: calc.js buildDealConfig now uses numOr() helper — blank/undefined falls back to defaults (25%/7%/22%) but a legitimate 0 is honored. Display layers fixed too (AssumptionsTab PITI preview + closing-costs panel, DealSummaryTab Financing KVs). TDD: src/lib/__tests__/calcAccuracy.test.js section A.",
     "The loan engine recommends 0%-down VA loans, but the calc engine could not model one — every VA deal was silently computed at 25% down.",
     "—"),

    ("BACK-096", "Financial Model",
     "Model PMI/MIP in cash flow, CoC, taxes, and FHA self-sufficiency",
     "P0 – Foundation", "Done (PROD)", "M",
     "As a low-down-payment buyer, I want my PMI to actually reduce my modeled cash flow so my returns aren't overstated.",
     "◇ Build complete: calc.js calcYear deducts pmi*12 from cashFlow/CoC/after-tax CF, deducts the rental share from taxable income, and auto-cancels when balance ≤ 78% of purchase price (HPA) or after a refi. NOT in NOI/DSCR (financing cost). FHA self-sufficiency PITI now includes MIP per HUD. Per-year pmi exposed in yearData; new '− PMI / MIP' row in CashFlowTab; PITI displays (DealSummaryTab hero/donut/Financing panel) include it. Sample deal now carries $150/mo PMI at 10% down. Glossary notes FHA life-of-loan caveat. Tests: calcAccuracy.test.js section B.",
     "PMI was collected in the UI and shown in PITI but ignored by every performance metric — $1,500–3,000/yr of phantom cash flow on low-down deals, including the onboarding sample deal.",
     "—"),

    ("BACK-097", "Financial Model",
     "Fix value-add remodel cost double-count in IRR / equity multiple",
     "P0 – Foundation", "Done (PROD)", "S",
     "As an investor modeling a renovation, I want the remodel cost counted once so my IRR isn't artificially crushed.",
     "◇ Build complete: calc.js calcExit — IRR t=0 outflow now uses base cash only (draws stay timed in yr1-2 cash flows); equity multiple adds draws back to the numerator. A $40k reno previously appeared as $80k of outflows. completionYear=1 now puts the full draw in yr 1. Tests: calcAccuracy.test.js sections C & I verify NPV(irr)≈0 on the explicit series.",
     "Every value-add deal materially understated IRR and equity multiple — the opposite bias from the rest of the audit, but still a wrong number.",
     "—"),

    ("BACK-098", "Financial Model",
     "Exit taxes on adjusted basis (§1250) + selling-cost assumption",
     "P0 – Foundation", "Done (PROD)", "M",
     "As an investor, I want exit proceeds to reflect selling costs and real depreciation-recapture tax so my IRR isn't systematically optimistic.",
     "◇ Build complete: calc.js calcExit — gain measured against adjusted basis (price + remodel draws − depreciation taken), amount realized nets out new sellingCostPct assumption (default 6%, range-clamped, editable in Assumptions → Growth & Exit). Selling at purchase price now correctly owes recapture. New fields sellingCosts/adjustedBasis returned; Exit panel (DealSummaryTab) and both exports show the selling-costs line. Legacy test asserting zero tax at zero appreciation rewritten — it encoded the bug. Tests: calcAccuracy.test.js section D.",
     "Net proceeds were overstated by ~6-8% commission PLUS understated exit tax (~$20k on a typical 10-yr hold) — every exit-dependent metric (IRR, EM, net proceeds, exit scenarios) had a structural rosy bias.",
     "—"),

    ("BACK-099", "Financial Model",
     "After-tax cash flow includes state & local tax",
     "P0 – Foundation", "Done (PROD)", "XS",
     "As a user in a taxed state, I want the After-Tax Cash Flow row to include the state tax shown directly above it.",
     "◇ Build complete: calc.js afterTaxCashFlow / afterTaxCFAdv now subtract totalStateTax (and PMI). CashFlowTab row tip updated. Tests: calcAccuracy.test.js section E.",
     "The table displayed a state-tax row and then an 'After-Tax' line that ignored it — internally inconsistent on the same screen.",
     "—"),

    ("BACK-100", "Financial Model",
     "§469 passive-loss limits in basic mode (PAL carryforward + sale release)",
     "P0 – Foundation", "Done (PROD)", "M",
     "As a high-AGI W-2 earner, I want paper losses limited per IRC §469 instead of getting an unlimited (illegal) deduction in the default mode.",
     "◇ Build complete: calc.js calcYear basic path applies the $25k active-participant allowance with the $100k–150k AGI phase-out (re_professional uncapped), suspends the excess into a basicPalCarryforward, applies it against future positive years, and releases it at sale (calcExit palTaxBenefit no longer adv-gated). CashFlowTab shows the carryforward derivation rows + PAL balance section in basic mode. Defaults (AGI 100k) preserve behavior for losses under $25k. Tests: calcAccuracy.test.js section F.",
     "Default mode granted every user full loss deductions against ordinary income — above $150k AGI the real answer is zero. Overstated after-tax cash flow for exactly the users most likely to buy.",
     "—"),

    # ─── 2026-06 Accuracy Audit — robustness (P2) ──────────────────────────
    ("BACK-101", "Financial Model",
     "IRR solver guardrails — bisection fallback + clamps",
     "P1 – Launch", "Done (PROD)", "S",
     "As a user with an unusual deal (refi cash-out, heavy losses), I want a finite, correct IRR instead of NaN or an absurd root.",
     "◇ Build complete: calc.js calcIRR — Newton's method with divergence/zero-derivative detection, verified against NPV, bisection fallback on [−99%, +1000%], returns 0 when there is no investment (cf0 ≥ 0). Tests: calcAccuracy.test.js section G.",
     "Refi cash-outs create multiple sign changes; Newton could converge to a wrong root or NaN and render garbage in the UI.",
     "—"),

    ("BACK-102", "Analysis",
     "Sensitivity fixes — listedRent fallback + vacancy clamp",
     "P1 – Launch", "Done (PROD)", "XS",
     "As a user whose rents come from listings, I want the Rent ±10% sensitivity to actually move my numbers.",
     "◇ Build complete: calc.js calcSensitivity perturbs effective rent (rent||listedRent) and clamps vacancy at 0% (no negative-vacancy income). Tests: calcAccuracy.test.js section H.",
     "Units priced via listedRent showed ZERO rent sensitivity — understating risk on exactly the deals where rent is least certain.",
     "—"),

    ("BACK-103", "Documents & Forms",
     "Export accuracy — PDF resolved-expense PITI, Loan Payoff field fix, selling-costs rows",
     "P1 – Launch", "Done (PROD)", "S",
     "As a user exporting a report, I want the PDF/XLSX numbers to match the app exactly.",
     "◇ Build complete: dealPDF.js PITI uses r.baseExpBreakdown (raw a.expenses was stale in %-of-rent mode); exit sidebar 'Loan Payoff' read years[].loanBalance which doesn't exist (field is .balance) — now r.exitLoanBalance; Selling Costs line added. dealXLSX.js exit section adds Selling Costs row and relabels tax row 'Tax on Sale (Recapture + LTCG)'; Net Proceeds KPI note updated.",
     "Exported PITI diverged from in-app numbers whenever expense %-mode was on, and the PDF Loan Payoff cell rendered from undefined.",
     "—"),

    ("BACK-104", "Financial Model",
     "Glossary accuracy updates from audit",
     "P2 – Growth", "Done (PROD)", "XS",
     "As a first-time buyer, I want tooltips that are technically correct so I learn the right things.",
     "◇ Build complete: glossary.js — pmi (cash-flow modeling + 78% cancel + FHA life-of-loan caveat), new sellingCosts entry, dscr (owner vs lender view), qbi (§199A safe-harbor caveat), sec1250Recapture (adjusted basis), effectiveMortgage (vacancy + PMI). Audit decisions documented: insurance-upfront in BOTH cash-to-close and yr-1 opex is intentional (prepaid convention); per-year capRate field is yield-on-cost but is not displayed per-year anywhere; breakEvenOccupancy treats %-of-rent expenses as fixed (known approximation).",
     "Education is the product's differentiator — the definitions must be right.",
     "—"),

    # ─── 2026-06 UX Audit ──────────────────────────────────────────────────
    ("UX-054", "UX & UI",
     "Assumptions Quick/Advanced split",
     "P1 – Launch", "Done (PROD)", "M",
     "As a new user, I want a short first-run form instead of ~50 fields so I can see my first analysis quickly.",
     "◇ Build complete: AssumptionsTab.jsx — advanced sections (Owner Occupancy, Growth & Exit, Tax Profile, Refinance, Value Add, Advanced Tax) sit behind one 'Show advanced settings' expander. Brand-new deals (no rents) default collapsed with a summary line of active defaults (OO status, growth, bracket, state, selling costs); deals with data default expanded. taxOpen hook hoisted to component top (Rules of Hooks — it previously lived inside the now-conditional block).",
     "First-run form drops from ~10 sections to 5; defaults are good enough to produce a credible Summary, and the summary line keeps hidden assumptions transparent.",
     "—"),

    ("UX-055", "UX & UI",
     "Deal Summary progressive disclosure + dead code removal",
     "P1 – Launch", "Done (PROD)", "S",
     "As a user, I want the Summary tab to lead with hero metrics, not a 6,000px wall of panels.",
     "◇ Build complete: DealSummaryTab.jsx — Exit Year Scenarios and Rentcast Property Data wrapped in CollapsibleSection (default collapsed; Exit badge shows selected year). Superseded chartData block (cumulative-by-multiplication bug, never rendered) deleted. egiExOO simplified (×12×(1−vac)÷12 → ×(1−vac)). Exit panel footnote now mentions selling costs + recapture.",
     "Cuts ~2 screens of below-fold content while keeping one-click access; removes a confusing dead-code trap.",
     "—"),

    ("UX-056", "Onboarding",
     "Guided tour — fix off-by-one tab indices, add Market step",
     "P1 – Launch", "Done (PROD)", "S",
     "As a new user taking the tour, I want each step to show the tab it describes.",
     "◇ Build complete: tourSteps.js — Showing/Red Flags/Sensitivity steps pointed at tabs 4/5/6 but the app's TAB_CONFIG has Market=4, Showing=5, Red Flags=6, Sensitivity=7; the tour narrated the wrong tabs. Indices fixed and a Market Data step added (tour grows 12→13 steps). Note: the earlier review claim that the 'advanced-features' anchor was missing was verified FALSE — AssumptionsTab has it and the expander button keeps it visible in quick mode.",
     "The onboarding tour literally showed the Market tab while describing Showing Notes — broken first impression.",
     "—"),

    ("UX-057", "Monetization",
     "Locked Share/PDF buttons open upgrade modal; tier-aware upgrade copy",
     "P1 – Launch", "Done (PROD)", "S",
     "As a locked user, I want clicking a Pro feature to show me how to upgrade, not blank the screen.",
     "◇ Build complete: DealPage.jsx — locked header buttons called setTab(-1), which rendered NO tab content (blank page). Now they open a dismissible overlay with UpgradeCard. UpgradeModal.jsx headline is tier-aware ('Unlock RentHack Pro' for locked users who may never have trialed; 'Go Pro before your trial ends' during trial) instead of always 'Your trial has ended'.",
     "The old behavior was worse than inert — it blanked the deal page. And mislabeled copy erodes trust at the exact moment of purchase intent.",
     "—"),

    ("UX-058", "UX & UI",
     "Expense $/% toggle hint + Loan Limit tooltip",
     "P2 – Growth", "Done (PROD)", "XS",
     "As a user, I want to discover the $/% expense mode and understand what Loan Limit does.",
     "◇ Build complete: AssumptionsTab.jsx — hint line at top of Expenses section explaining the $/% toggle; Tip on Loan Limit label explaining the cap behavior and the extra-down-payment consequence.",
     "Audit found most users would never discover %-of-rent mode and couldn't tell why Loan Amount wasn't editable.",
     "—"),

    # ─── Deferred from audit (not shipped this round) ──────────────────────
    ("BACK-105", "Infrastructure",
     "Excel projection-formula parity tests",
     "P2 – Growth", "Backlog", "M",
     "As a developer, I want automated tests proving the XLSX projection formulas (Years 2+) match calc.js so a future calc change can't silently desync the export.",
     "From 2026-06 audit: export.test.js is smoke-only. Needs a harness that evaluates the generated Excel formulas (e.g. via a formula parser) against calc.js outputs for a fixture deal. Deferred — meaningful tooling work.",
     "An export that silently diverges from the app is an accuracy time bomb.",
     "BACK-103"),

    ("BACK-106", "Monetization",
     "Server-side tier enforcement for gated features",
     "P2 – Growth", "Backlog", "L",
     "As the owner, I want Pro gating enforced server-side so a technical user can't bypass the paywall by editing client state.",
     "From 2026-06 audit: computeTier() reads user_metadata client-side; exports run fully in-browser so gating is honor-system. Options: move export generation behind an edge function, or verify entitlements from a server-checked table on feature use. Stripe payment itself is already server-side.",
     "Client-only gating means the paywall is advisory for anyone who opens devtools.",
     "—"),

    ("BACK-107", "Infrastructure",
     "Persistent offline sync queue (IndexedDB)",
     "P3 – Future", "Backlog", "L",
     "As a user who edits a deal and immediately closes the tab, I don't want to lose the edit in the 800ms debounce window.",
     "From 2026-06 audit: pendingDealIds lives in a useRef and dies on crash/refresh; the localStorage↔Supabase bootstrap ordering is correct but protected only by a comment. A persisted write queue (IndexedDB) with replay on boot would harden it.",
     "Rare but real data-loss window; also de-fragilizes the bootstrap ordering constraint.",
     "—"),

    ("UX-059", "UX & UI",
     "Tab consolidation 9 → 7 (Comps+Market, Showing+Red Flags)",
     "P3 – Future", "Needs Design", "L",
     "As a user, I want fewer, clearer tabs: Analyze → Market → Due Diligence → Decide.",
     "From 2026-06 audit: merge Rent Comps into Market ('Market' with comps section) and Showing into Red Flags ('Due Diligence'). Cuts cognitive load with no functionality loss. Needs design pass on gating (rentComps gate applies to a sub-section) and mobile tab labels before build.",
     "9 tabs is the single biggest structural driver of the 'overwhelming' feedback.",
     "—"),
]

COLS = 10

def main():
    if not SRC.exists():
        print(f"ERROR: {SRC} not found"); sys.exit(1)
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["Product Backlog"]

    # Build ID → row map
    id_row = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v: id_row[str(v)] = r

    for (rid, status, note) in STATUS_UPDATES:
        if rid not in id_row:
            print(f"  !! {rid} not found, skipping"); continue
        r = id_row[rid]
        ws.cell(row=r, column=5).value = status
        if note:
            cur = ws.cell(row=r, column=8).value or ""
            ws.cell(row=r, column=8).value = (cur + "\n◇ " + note).strip()
        print(f"  updated {rid} -> {status}")

    # Append new rows, copying style from the last existing data row
    template_row = ws.max_row
    for new in NEW_ROWS:
        r = ws.max_row + 1
        for c in range(1, COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = new[c - 1]
            src_cell = ws.cell(row=template_row, column=c)
            cell.font = copy(src_cell.font)
            cell.fill = copy(src_cell.fill)
            cell.border = copy(src_cell.border)
            cell.alignment = copy(src_cell.alignment)
        print(f"  appended {new[0]}: {new[2][:60]}")

    wb.save(DST)
    print(f"\nSaved {DST.name}: {len(STATUS_UPDATES)} updates, {len(NEW_ROWS)} new rows.")
    print("Reminder: upload to the RentHack Google Drive folder and bump CLAUDE.md refs.")

if __name__ == "__main__":
    main()
