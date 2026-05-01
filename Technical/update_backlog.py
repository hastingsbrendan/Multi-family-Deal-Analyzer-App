"""
Update RentHack_Product_Backlog_v15.xlsx with the work shipped across Sprints 1-6
plus the Batch 1-5 polish round, then save as v17.

Two operations:
  1) Mark existing Backlog items as Done (PROD) where shipped.
  2) Append new rows for work that didn't have an existing ID.

Run from project root:
  python Technical/update_backlog.py
"""
import sys, io, shutil
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import PatternFill, Alignment
from copy import copy
from pathlib import Path

PROJ = Path(__file__).parent.parent
SRC = PROJ / "RentHack_Product_Backlog_v15.xlsx"
DST = PROJ / "RentHack_Product_Backlog_v17.xlsx"

# ─── 1. Status updates for existing rows ────────────────────────────────────
STATUS_UPDATES = [
    # (ID, new_status, optional updated AC note appended with ' — ' separator)
    ("803", "Done (PROD)",
     "Shipped Sprint 2: ComparePanel in PortfolioPage with checkbox selection (max 3 deals), side-by-side metric comparison covering Year 1 performance, hold-period returns, and deal info. Reuses cached calcDeal results."),

    ("910", "Done (PROD)",
     "Shipped Sprint 6: For OO deals, Effective Mortgage rendered as the FIRST hero card on Deal Summary (above Profitability section). Plain-English subline explains tenant rent coverage, plus side metrics for PITI / Tenant Rent / vs-Renting delta. Violet --oo-violet accent border for house-hack signal."),

    ("920", "Done (PROD)",
     "Shipped Sprint 6: 'What if I move out?' pill toggle for OO deals on Deal Summary. When on, runs calcDeal a second time with ownerOccupied=false and shows side-by-side Year 1 Cash Flow / Cap Rate / CoC / IRR. Memoized via useMemo([deal])."),

    ("UX-017", "Done (PROD)",
     "Shipped Sprint 6: Inside a deal on mobile, header hides lastSyncedAt timestamp + feedback button + refresh button. Those actions move into the avatar dropdown ('Send Feedback' / 'Refresh from Cloud'). Header keeps brand+breadcrumb, sync badge, PRO pill, avatar."),

    ("UX-018", "Done (PROD)",
     "Shipped Batch 4: Mobile Cash Flow tab now defaults to a compact view (Yr1 + Exit Year only). Toggle pill above the table expands to all hold-period years. Right-edge scroll shadow that fades when scrolled to end. Removes the 'swipe through 10 narrow columns' friction on a 360px viewport."),
]

# ─── 2. New rows to append ──────────────────────────────────────────────────
# Each row maps to the 10 Product Backlog columns:
#   ID, Epic, Feature/Task, Priority, Status, Effort, User Story,
#   Acceptance Criteria/Build Notes, Why It Matters, Dependencies
NEW_ROWS = [
    # ─── Sprint 1 — Foundations (design system, tooltips, empty/sample) ────
    ("UX-043", "UX & UI",
     "Design tokens — type/spacing/radius/shadow/transition scales in index.css",
     "P1 – Launch", "Done (PROD)", "M",
     "As a developer, I want a single source of truth for type/spacing/radius/shadow values so components stay visually consistent and theme-aware.",
     "Added to :root in src/styles/index.css: --text-{xs,sm,base,md,lg,xl,2xl} (Major Third 1.25 scale), --sp-{1..8} (4px grid), --r-{sm,md,lg,xl,pill} radius scale, --shadow-{sm,md,lg,xl} elevation, --t-fast/--t-base transitions. Theme-agnostic so both light + dark inherit them.",
     "Foundation for the design-system migration. Without tokens, every component picks its own fontSize/borderRadius and visual inconsistency creeps back.",
     "—"),

    ("UX-044", "UX & UI",
     "ui/ primitives — Panel, KeyValue, Pill, Button, EmptyState, SectionHeader, Tip",
     "P1 – Launch", "Done (PROD)", "L",
     "As a developer, I want reusable React primitives for buttons/panels/pills/etc. so I stop reimplementing the same inline styles in every component.",
     "Created src/components/ui/: Panel.jsx (card container with optional accent + elevation), KeyValue.jsx (label-value pair with optional bottom border), Pill.jsx (compact rounded label with 7 variants), Button.jsx (5 variants × 3 sizes), EmptyState.jsx (icon + title + body + CTAs), SectionHeader.jsx, Tip.jsx (glossary tooltip with viewport-edge clamping).",
     "DRYs out hundreds of duplicated style blocks across components. Enables future redesigns by changing primitive once.",
     "UX-043"),

    ("UX-045", "UX & UI",
     "Glossary tooltips — ~30 Tip placements across Deal Summary, Cash Flow, Red Flags",
     "P1 – Launch", "Done (PROD)", "M",
     "As a novice investor, I want plain-English explanations next to investor jargon so I can learn the terminology without leaving the app.",
     "src/lib/glossary.js holds tooltip definitions for DSCR, NOI, Cap Rate, IRR, CoC, vacancy rate, etc. ~30 <Tip text={GLOSSARY.X}/> placements added to DealSummaryTab, CashFlowTab, RedFlagsTab on every metric label.",
     "Removes the #1 barrier for first-time users (terminology). Cost: low. Impact: every novice gets a built-in glossary.",
     "UX-044"),

    ("UX-046", "UX & UI",
     "New deals open on Assumptions tab when empty (no rents/price)",
     "P1 – Launch", "Done (PROD)", "S",
     "As a new user creating my first deal, I want to land on the input form rather than an empty Deal Summary so I know what to do next.",
     "DealPage.jsx initial tab logic: if a deal has purchasePrice===0 AND all units have zero rent, open on tab 1 (Assumptions) instead of tab 0 (Summary). Otherwise default to Summary as before.",
     "Eliminates first-deal confusion ('why is everything empty?'). One-line fix, big UX impact.",
     "—"),

    ("UX-047", "UX & UI",
     "Sample deal CTA on empty portfolio page",
     "P1 – Launch", "Done (PROD)", "S",
     "As a first-touch user, I want to see a working analysis immediately instead of a blank form, so I can evaluate whether RentHack is for me before doing data entry.",
     "PortfolioPage empty-state EmptyState now exposes secondary={label:'Try a sample deal →', onClick: onAddSample}. Calls createSampleDeal() from calc.js (fully-loaded Chicago duplex). The _isSample flag distinguishes sample deals in the portfolio.",
     "Better activation funnel: instead of 'No deals yet', new users see a working deal in 1 click and can poke around.",
     "—"),

    # ─── Sprint 2 — UX wins ───────────────────────────────────────────────
    ("UX-048", "UX & UI",
     "Settings consolidation — single tabbed page (Defaults | Account | Appearance | Groups)",
     "P1 – Launch", "Done (PROD)", "M",
     "As a user, I want a single Settings entry point instead of 3 separate menu items so I don't have to remember which one has what.",
     "Replaced 3 separate menu items (Default Deal Settings / Settings / My Profile) with a single SettingsPage with 4 tabs. Avatar-menu shows just '⚙️ Settings'. Tab bar: Defaults (push-to-existing field selector), Account (profile + password + data export + delete account), Appearance (dark mode), Groups (CTA into Groups page).",
     "Reduces nav cognitive load. Single discoverable Settings location for everything.",
     "UX-044"),

    ("UX-049", "UX & UI",
     "Fraunces serif on hero metrics; removed 47 system-ui fontFamily overrides",
     "P2 – Growth", "Done (PROD)", "S",
     "As a user, I want the app's typography to feel premium and consistent so it matches the editorial style RentHack is going for.",
     "DealSummaryTab: 5 hero metrics now use fontFamily:\"'Fraunces',serif\" (Avg Monthly Equity Growth, IRR, Cap Rate, EM, CF). Removed 47 inline fontFamily:\"system-ui\" overrides that were hijacking the global DM Sans inheritance.",
     "Brand voice. Mockups show Fraunces hero numbers; previous code muted them with system-ui everywhere.",
     "UX-043"),

    # ─── Sprint 3 — Refactor (calc decomposition + tests + extraction) ────
    ("BACK-073", "Infrastructure",
     "calcDeal decomposition — 4 internal helpers (calcIRR, buildDealConfig, calcYear, calcExit)",
     "P1 – Launch", "Done (PROD)", "L",
     "As a developer, I want calcDeal split into testable sub-functions so I can target tax/operating logic in isolation and unblock future scenario features.",
     "src/lib/calc.js: calcDeal was a ~300-line monolith. Split into 4 internal (non-exported) helpers: calcIRR (Newton-Raphson), buildDealConfig (loan + OO + refi + VA + tax setup), calcYear (per-year operating + tax with threaded state), calcExit (§1250 recapture + IRR + EM + FHA self-suff). calcDeal becomes a ~40-line orchestrator. All 73 prior tests pass unchanged.",
     "Unblocks BACK-803 (deal comparison) and future portfolio dashboard work — both want to call sub-pieces, not the whole engine. Also makes the highest-risk code (advanced tax + PAL) testable in isolation.",
     "—"),

    ("BACK-074", "Infrastructure",
     "Test coverage — taxEngine (32), loanEngine (19), floodZone (13) unit tests",
     "P1 – Launch", "Done (PROD)", "M",
     "As a developer, I want unit tests on the 50-state tax engine and loan recommendation engine so silent data drift gets caught.",
     "Created src/lib/__tests__/taxEngine.test.js (32 tests: no-tax states, flat rates, progressive bracket stacking, filing status, local tax, edge cases, bracket integrity sweep). loanEngine.test.js (19 tests: catalog integrity, OO+veteran VA path, OO+excellent-credit Conv, low-credit FHA, jumbo detection, question flow). floodZone.test.js (13 tests: every zone classification including aliases).",
     "Bracket data drift is invisible without tests. Loan engine had 472 known no-recommendation scenarios per audit — tests pin the contract before any LOAN-003 fixes.",
     "—"),

    ("BACK-075", "Infrastructure",
     "AssumptionsTab extraction — FmtInt, CollapsibleSection, PropertyLookupPanel to own files",
     "P2 – Growth", "Done (PROD)", "M",
     "As a developer, I want the 1357-line AssumptionsTab split so the file is easier to read and the helpers are reusable elsewhere.",
     "Extracted to own files: FmtInt → src/components/ui/FmtInt.jsx, CollapsibleSection → src/components/ui/CollapsibleSection.jsx, PropertyLookupPanel + parseAddressFromUrl → src/components/AssumptionsTab/PropertyLookupPanel.jsx. Section bodies (Basics, Financing, etc.) stay inline (heavy upd() closure coupling — extraction ROI low).",
     "AssumptionsTab dropped 1357 → 1002 lines. Extracted primitives ready to use elsewhere.",
     "—"),

    ("BACK-076", "Infrastructure",
     "MarketTab extraction — helpers, RateContextPanel, AssumptionsCheckPanel, MarketUIHelpers",
     "P2 – Growth", "Done (PROD)", "M",
     "As a developer, I want MarketTab's data-fetching helpers and standalone panels split out so the main component is focused on data orchestration.",
     "Created src/components/MarketTab/: marketHelpers.js (CENSUS_VARS, FRED_BATCH, BLS/QCEW helpers, parseFredObs/yoyPct/etc.), MarketUIHelpers.jsx (SectionHeader, StatRow, BenchmarkRow, MktSection, MktEmptyState, ChartTooltip, RateCompare), RateContextPanel.jsx, AssumptionsCheckPanel.jsx.",
     "MarketTab dropped 1096 → 803 lines. Helpers usable in future market-data features.",
     "—"),

    ("BACK-077", "Infrastructure",
     "DealSummaryTab — hoisted SubHdr/SLbl/KV/DSPanel from render to module scope",
     "P2 – Growth", "Done (PROD)", "S",
     "As a developer, I want the inline UI helpers in DealSummaryTab hoisted so they're not recreated on every render.",
     "SubHdr, SLbl, KV, DSPanel previously defined inside the DealSummaryTab function body (recreated on every render, not reusable). Hoisted to module scope. DSPanel now wraps the shared <Panel> primitive (preserving 12px radius for visual continuity).",
     "Eliminates per-render allocation. Routes the panel through the design system.",
     "UX-044"),

    ("BACK-078", "Infrastructure",
     "PortfolioPage — memoize calcDeal per deal on _deal_id+updated_at",
     "P2 – Growth", "Done (PROD)", "S",
     "As a user with a portfolio of 10+ deals, I want the portfolio page to not re-run calcDeal for every deal on every render.",
     "PortfolioPage uses a useRef-backed cache keyed on _deal_id+updated_at. resultsByDealId useMemo only recalculates deals whose updated_at changed; unchanged deals reuse the cached calcDeal result.",
     "Avoids N×calcDeal per render. Noticeable on portfolios with many deals.",
     "—"),

    ("BACK-079", "Infrastructure",
     "Hex audit — replaced #dc2626 → var(--red); rest intentional",
     "P2 – Growth", "Done (PROD)", "S",
     "As a developer, I want any remaining hardcoded hex colors that map to existing CSS vars replaced so the design system is the single source of color truth.",
     "DealSummaryTab: 2× '#dc2626' → 'var(--red)'. Remaining hex values are intentional: chart palettes (need distinct colors, not theme-aware), Google SVG brand colors (must not change), alpha-appended hex like STATUS_COLORS+'22' (handled separately in BACK-088), fixed-dark toast styles.",
     "Closes the easy wins. The remaining 60 hex values aren't actually theme-bugs.",
     "UX-043"),

    # ─── Sprint 5 — Refactor (LoanType + export) ──────────────────────────
    ("BACK-080", "Infrastructure",
     "LoanTypeTab extraction — ScoreBadge, ProgressBar, QuestionCard moved out",
     "P2 – Growth", "Done (PROD)", "M",
     "As a developer, I want LoanTypeTab's 838-line file split so the loan-quiz UX is easier to navigate.",
     "Extracted: ScoreBadge → src/components/ui/ScoreBadge.jsx (5-tier loan match badge), ProgressBar → src/components/ui/ProgressBar.jsx (3px slim progress indicator), QuestionCard → src/components/LoanTypeTab/QuestionCard.jsx (~115 lines, both choice + slider question types). Removed dead parseStyleString helper.",
     "LoanTypeTab dropped 838 → 689 lines. ProgressBar reusable elsewhere (e.g. trial banner).",
     "—"),

    ("BACK-081", "Infrastructure",
     "lib/export/ split — colors, helpers, portfolioXLSX, dealXLSX, dealPDF, index",
     "P1 – Launch", "Done (PROD)", "L",
     "As a developer, I want export.js's 1663 lines split into focused files since this is the most critical untested code path (paid feature).",
     "Created src/lib/export/: colors.js (shared XLSX hex + jsPDF RGB palette), helpers.js (dlFile blob downloader), portfolioXLSX.js (305 lines), dealXLSX.js (745 lines), dealPDF.js (579 lines), index.js (public re-export). src/lib/export.js → 10-line backward-compat shim.",
     "Each export type now lives in its own file. Easier to maintain when jspdf or xlsx-js-style upgrades break things.",
     "—"),

    ("BACK-082", "Infrastructure",
     "export.js smoke tests — 8 tests across all 3 export pipelines",
     "P1 – Launch", "Done (PROD)", "M",
     "As a developer, I want smoke tests on the export pipeline so dependency upgrades (jspdf, xlsx-js-style) can't silently break paid features.",
     "src/lib/__tests__/export.test.js: 3 portfolioXLSX scenarios (full, empty, null user / CSV path), 2 dealXLSX (basic + owner-occupied), 3 dealPDF (basic + no-address + refi-enabled). jsPDF + jspdf-autotable mocked to avoid filesystem writes during tests. Total tests: 172 → 180.",
     "Catches regressions on dep upgrades. PDF/Excel exports are revenue-critical (rent-comps + sensitivity gated; export is the receipt for trial users).",
     "BACK-081"),

    # ─── Sprint 6 — UX wins (additions beyond UX-017/018, 910, 920) ───────
    ("UX-050", "UX & UI",
     "Wire sample-deal CTA → guided tour auto-fire (gated by per-user localStorage flag)",
     "P1 – Launch", "Done (PROD)", "S",
     "As a first-touch user, I want the guided tour to auto-launch when I click 'Try a sample deal' so I don't have to find it separately.",
     "addSampleDeal callback in App.jsx now also calls setTourStep(0) for first-touch users. Gated by localStorage 'rh_tour_seen_<userId>' flag so it doesn't re-fire on later sample additions.",
     "Activation funnel: sample-deal CTA + tour together produce a complete first-run, instead of dropping the user into a sample with no guidance.",
     "UX-047"),

    ("UX-051", "UX & UI",
     "Sync visibility — one-time toast for new users after first cloud save",
     "P2 – Growth", "Done (PROD)", "M",
     "As a new user, I want explicit confirmation that my work is being saved to the cloud, since the existing tiny ✓ Synced badge is easy to miss.",
     "App.jsx useEffect on syncStatus==='saved' && !localStorage[`rh_sync_toast_shown_${user.id}`]: shows 'Your work is saved automatically. Close anytime.' banner for 5s, then sets the flag so it doesn't re-fire.",
     "Reduces 'is it saving?' anxiety for first-touch users without spamming returning users.",
     "—"),

    ("UX-052", "UX & UI",
     "PRO pill in nav for paid users",
     "P2 – Growth", "Done (PROD)", "S",
     "As a paid Pro subscriber, I want a small visual recognition that I'm a paying customer.",
     "App.jsx nav header: when useSubscription().tier === 'pro', renders <Pill variant='accent' size='xs'>PRO</Pill> next to the avatar.",
     "Recognition matters for retention. Trivial cost.",
     "MON-001"),

    # ─── Batch 1-5 — Polish round (May 2026) ──────────────────────────────
    ("BACK-083", "Infrastructure",
     "structuredClone migration — 27 JSON.parse(JSON.stringify) calls replaced",
     "P2 – Growth", "Done (PROD)", "S",
     "As a developer, I want native structuredClone() instead of the JSON-roundtrip pattern since it's faster and supports Date/Map/Set without surprises.",
     "Bulk sed replacement across 10 files: src/lib/calc.js (3), src/components/{App, AssumptionsTab/PropertyLookupPanel, AssumptionsTab, DealSummaryTab, MarketTab, RedFlagsTab, RentCompsTab, SensitivityTab, ShowingTab}.jsx. Convention note in CLAUDE.md updated.",
     "Performance + correctness. Modern browsers and Node 17+ all support it.",
     "—"),

    ("BACK-084", "UX & UI",
     "CSS vars: --warn-bg/--warn-text and --status-{state}-bg",
     "P2 – Growth", "Done (PROD)", "S",
     "As a developer, I want semantic CSS vars for the amber 'warning' pattern and per-status pill backgrounds so theme-switching adjusts the alpha tint correctly.",
     "Added to both [data-theme=light] and [data-theme=dark] in src/styles/index.css: --warn-bg / --warn-text (amber pair), --status-{analyzing,under-contract,owned,pass}-bg (pill alpha tints). Light/dark have different alpha to keep contrast right in dark mode.",
     "Replaces ad-hoc 'STATUS_COLORS+\"22\"' alpha-hex concatenation. Sets up BACK-088 status pill migration.",
     "UX-043"),

    ("UX-053", "UX & UI",
     "ui/Spinner primitive — replaces 4 ad-hoc 'Loading…' snippets",
     "P2 – Growth", "Done (PROD)", "S",
     "As a developer, I want a single Spinner component so loading states look consistent across the app.",
     "Created src/components/ui/Spinner.jsx with 3 modes: inline (in-flow with text), block (default, padded box), fullPage (vh-100 centered). Adopted in App.jsx (3 places: deals===null, authLoading, deals undefined fallback) and DealPage.jsx (TabFallback). rh-spin keyframe added to index.css.",
     "Visual consistency. Easier to swap loading affordance later (e.g. add skeletons).",
     "UX-044"),

    ("BACK-085", "Infrastructure",
     "validateDealShape now runs at sbWriteDeal boundary",
     "P2 – Growth", "Done (PROD)", "S",
     "As a developer, I want malformed deals caught before they're persisted to the cloud, not just on read.",
     "Previously validateDealShape() was only called in loadLocal + sbRead. Added a third call site in sbWriteDeal (constants.js) so any drift between in-memory shape and round-tripped shape is captured to Sentry. Non-blocking — write proceeds even if shape is invalid.",
     "Closes the read-only validation gap. Useful when a future migration changes deal shape but a stale client writes the old shape.",
     "—"),

    ("BACK-086", "Infrastructure",
     "groups.js test coverage — 20 smoke tests",
     "P2 – Growth", "Done (PROD)", "M",
     "As a developer, I want test coverage on the groups/comments Supabase API since 252 lines of mutations had zero tests.",
     "src/lib/__tests__/groups.test.js (209 lines, 20 tests): chainable Supabase client mock that records every method call on the query builder. Tests verify the right table is queried and the right mutation type runs for all 17 exported functions (sbGetMyGroups, sbCreateGroup, sbInviteMember, sbRespondToInvite, comments CRUD, etc.).",
     "Group sharing is a paid feature; auth bugs would be invisible until customer reported. Total tests: 180 → 200.",
     "—"),

    ("BACK-087", "UX & UI",
     "SettingsPage — extracted Appearance + Groups tabs to own files",
     "P3 – Future", "Done (PROD)", "S",
     "As a developer, I want SettingsPage's smaller tabs split out so the main file is more navigable.",
     "Created src/components/SettingsPage/: AppearanceTab.jsx (dark mode + legal footer), GroupsTab.jsx (CTA into Groups page). Defaults + Account stay inline due to heavy state coupling (15+ pieces of state) — extraction ROI low for those, tracked separately as BACK-092.",
     "Establishes the SettingsPage/ directory pattern for future tab additions.",
     "UX-048"),

    ("BACK-088", "UX & UI",
     "Status pill alpha-hex migration — STATUS_BG_VARS replaces STATUS_COLORS+'22'",
     "P2 – Growth", "Done (PROD)", "S",
     "As a developer, I want status pill backgrounds to come from CSS vars so they're theme-aware and don't rely on alpha-hex string concatenation.",
     "Added STATUS_BG_VARS map in constants.js mapping each status to its --status-{state}-bg CSS var. 4 inline call sites updated: PortfolioPage card view + table view, DealPage status select, ComparePanel.",
     "Theme-aware status tints (dark mode gets a slightly higher alpha automatically). Removes string concatenation hack.",
     "BACK-084"),

    ("BACK-089", "UX & UI",
     "Sprint 4 design-system primitive adoption — ~25 buttons + DSPanel/MktSection + Sample/PRO badges",
     "P1 – Launch", "Done (PROD)", "M",
     "As a developer, I want the ui/ primitives actually used by components so the Sprint 1 foundation pays off.",
     "Across 10 files (PortfolioPage, RedFlagsTab, FeedbackModal, ProfilePage, DealPage, CommentsPanel, UpgradeModal, ShareDealModal, GuidedTour, SettingsPage): ~25 raw <button> tags → <Button variant=...>. DealSummaryTab DSPanel and MarketTab MktSection now wrap <Panel>. App.jsx PRO badge + PortfolioPage Sample badge → <Pill variant='accent'>. Replaced ~98 lines of duplicated inline button styles.",
     "Pays off the Sprint 1 ui/ work. Future redesigns change the primitive once instead of every call site.",
     "UX-044"),

    ("BACK-090", "UX & UI",
     "Sprint 4d token adoption — 140 inline values migrated to tokens in DealSummaryTab + AssumptionsTab",
     "P2 – Growth", "Done (PROD)", "M",
     "As a developer, I want the design tokens (--text-*, --r-*) actually used in the largest tabs so visual consistency stops drifting.",
     "DealSummaryTab + AssumptionsTab: 29 borderRadius:N → var(--r-{sm,md,lg,xl,pill}); 111 fontSize:N → var(--text-{xs,sm,base,md,lg,xl,2xl}). Untouched: numbers that don't map to tokens (radius 4/8/12; fontSize 9/10/12/15/17/18/22/24/26/32/36/44).",
     "Closes the gap noted in the round-2 review: tokens were defined but barely consumed. Visual consistency now has a working contract.",
     "UX-043"),

    # ─── New backlog items (deferred work + follow-ups) ───────────────────
    ("BACK-091", "Infrastructure",
     "Test infrastructure — install jsdom + @testing-library/react",
     "P2 – Growth", "Backlog", "M",
     "As a developer, I want React component test infrastructure so I can write smoke tests that protect gated-feature/BlurGate logic and useCloudSync state machines.",
     "npm install -D jsdom @testing-library/react @testing-library/jest-dom. Add `environment: 'jsdom'` to vite.config.js test block. With this in place, write: useCloudSync state-machine tests (mock Supabase client, assert syncStatus transitions through saving/saved/offline/error), component smoke tests for DealPage gated tabs (rent comps, sensitivity), and BlurGate visibility for tier=='locked'.",
     "Subscription tier enforcement is revenue-critical. Without component tests, gating logic could break silently between deploys.",
     "—"),

    ("BACK-092", "Infrastructure",
     "SettingsPage — extract Defaults + Account tabs (heavy state coupling, low ROI)",
     "P3 – Future", "Backlog", "M",
     "As a developer, I want SettingsPage fully split so the parent component is just a tab-bar shell.",
     "Defaults tab uses 6 pieces of state (local, saving, saved, pushFields, showPushPanel, plus PUSH_FIELDS/SECTIONS data). Account tab uses 15+ pieces (displayName, organization, profSaving, profSaved, profErr, changingPw, newPw, confirmPw, pwSaved, pwErr, exportLoading, exportDone, showDeleteModal, deleteConfirm, deleteLoading, deleteError, plus saveProfile/changePw/handleExport/handleDeleteAccount). Either pass everything via props or move state into the tab component itself.",
     "Currently SettingsPage is 495 lines after Appearance+Groups extraction. Defaults+Account split would drop it further but the prop-passing or state-internalization adds friction.",
     "BACK-087"),

    ("BACK-093", "UX & UI",
     "Continue Button primitive migration — ~150 raw <button> remain in big tabs",
     "P2 – Growth", "Backlog", "M",
     "As a developer, I want the remaining inline buttons migrated to <Button> so the design-system migration is complete.",
     "After BACK-089, ~150 raw <button> tags remain — mostly in: AssumptionsTab body (input mode toggles, OO unit selectors), MarketTab body (filter pills, retry buttons), LoanTypeTab body (option cards, slider continue), GroupsPage (member management actions), AuthScreen (Google button has unique styling — keep inline). Many are icon-only or have conditional dynamic styling that doesn't fit Button variants cleanly.",
     "Finish the migration so visual inconsistency doesn't creep back. Some buttons may stay inline if Button's variants don't fit.",
     "BACK-089"),

    ("BACK-094", "Infrastructure",
     "useCloudSync state-machine tests (depends on BACK-091)",
     "P2 – Growth", "Backlog", "M",
     "As a developer, I want regression coverage on the cloud-sync state machine since the code has the most heavily-commented race-condition fixes in the codebase.",
     "After BACK-091 adds jsdom + RTL, write: idle→saving→saved transition; saving→error on Supabase failure; offline→saved when isOnline flips back; granular vs bulk write paths (dirty.size>0 vs ===0); _deal_id back-fill on first write of new deal; debounced timer cancellation on rapid edits.",
     "Without tests, the next race-condition fix could regress silently. The 800ms debounce + dirty-tracking pattern is subtle.",
     "BACK-091"),
]

# ─── Apply updates ────────────────────────────────────────────────────────
def main():
    shutil.copy(SRC, DST)
    wb = openpyxl.load_workbook(DST, data_only=False)
    ws = wb['Product Backlog']

    # Build ID -> row index lookup
    id_to_row = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0]:
            id_to_row[str(row[0])] = i

    print("=== Status updates ===")
    for rid, new_status, ac_append in STATUS_UPDATES:
        r = id_to_row.get(rid)
        if not r:
            print(f"  SKIP {rid}: not found")
            continue
        prev_status = ws.cell(row=r, column=5).value
        ws.cell(row=r, column=5).value = new_status
        # Append to AC column (8) as a "Build complete:" addendum
        prev_ac = ws.cell(row=r, column=8).value or ""
        if "Build complete:" not in prev_ac:
            ws.cell(row=r, column=8).value = (prev_ac.rstrip() + "  ◇ Build complete: " + ac_append).strip()
        print(f"  {rid}: {prev_status!r} -> {new_status!r}")

    # Find first empty row to start appending
    append_row = ws.max_row + 1
    print(f"\n=== Appending {len(NEW_ROWS)} new rows starting at row {append_row} ===")

    # Determine the style to clone from a representative existing data row
    # so new rows visually match (font, alignment, wrap).
    template_row = 5
    template_cells = [ws.cell(row=template_row, column=c+1) for c in range(10)]

    for offset, row_data in enumerate(NEW_ROWS):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=append_row + offset, column=c)
            cell.value = val
            tpl = template_cells[c - 1]
            cell.font = copy(tpl.font)
            cell.alignment = copy(tpl.alignment) if tpl.alignment else Alignment(wrap_text=True, vertical="top")
            cell.border = copy(tpl.border)
            cell.fill = copy(tpl.fill)
        print(f"  + {row_data[0]} ({row_data[4]}) — {row_data[2][:60]}")

    wb.save(DST)
    print(f"\nSaved → {DST}")
    print(f"Total rows now: {ws.max_row}")

if __name__ == "__main__":
    main()
